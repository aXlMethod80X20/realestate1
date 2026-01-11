"""
Norway Hotel Database - Full Scraper
Discovers AND enriches all hotels/B&Bs in Norway from scratch.

Sources:
1. Brreg API (free) - Find all hotel companies by NACE code
2. Google Places API (free 300/day) - Commercial names, ratings, address, phone, website
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
import pandas as pd
from datetime import datetime
import os
import time
import re
import random

# ============================================================
# CONFIGURATION
# ============================================================
# Set your API key via environment variable or paste here
GOOGLE_PLACES_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")

# Fylkesnummer prefix for filtering (first 2 digits of kommunenummer)
FYLKE_PREFIX = {
    "Nord-Norge": ["18", "54", "55"],  # Nordland (18xx), Troms (54xx), Finnmark (55xx)
    "Trøndelag": ["50"],  # Trøndelag (50xx)
    "Vestlandet": ["11", "46", "15"],  # Rogaland (11xx), Vestland (46xx), Møre og Romsdal (15xx)
    "Østlandet": ["03", "31", "32", "33", "34", "38", "39", "40"],  # Oslo, Viken, Innlandet, etc
    "Sørlandet": ["42"],  # Agder (42xx)
    "Hele Norge": []  # All - no filter
}

# NACE codes for accommodation (use short codes - API returns 55.100, 55.200, etc)
NACE_CODES = {
    "hotels": ["55.100"],  # Hotels
    "bb": ["55.200", "55.900"],  # B&B, hostels, other accommodation
    "camping": ["55.300"]  # Camping
}


class HotelScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Norway Hotel Database - Full Scraper")
        self.root.geometry("1200x700")

        self.hotels = []
        self.is_running = False
        self.api_calls = 0  # Track Google API calls
        self.MAX_API_CALLS = 300  # Free limit

        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)

        # Title
        ttk.Label(main_frame, text="Norway Hotel Database", font=('Helvetica', 16, 'bold')).grid(row=0, column=0, pady=(0, 5))
        ttk.Label(main_frame, text="Discovery: Brreg API (free) | Enrichment: Google Places (300/day free)", font=('Helvetica', 9)).grid(row=1, column=0, pady=(0, 15))

        # Settings
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="10")
        settings_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))

        ttk.Label(settings_frame, text="Region:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.region_var = tk.StringVar(value="Nord-Norge")
        ttk.Combobox(settings_frame, textvariable=self.region_var, values=list(FYLKE_PREFIX.keys()), width=20).grid(row=0, column=1, sticky="w")

        ttk.Label(settings_frame, text="Include:").grid(row=0, column=2, sticky="w", padx=(20, 10))
        self.include_hotels = tk.BooleanVar(value=True)
        self.include_bb = tk.BooleanVar(value=True)
        self.include_camping = tk.BooleanVar(value=False)
        ttk.Checkbutton(settings_frame, text="Hotels", variable=self.include_hotels).grid(row=0, column=3)
        ttk.Checkbutton(settings_frame, text="B&B/Pensjonat", variable=self.include_bb).grid(row=0, column=4)
        ttk.Checkbutton(settings_frame, text="Camping", variable=self.include_camping).grid(row=0, column=5)

        ttk.Label(settings_frame, text="Max hotels:").grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.limit_var = tk.StringVar(value="300")
        ttk.Entry(settings_frame, textvariable=self.limit_var, width=10).grid(row=1, column=1, sticky="w", pady=(10, 0))
        ttk.Label(settings_frame, text="(max 300 for free Google enrichment)", font=('Helvetica', 8)).grid(row=1, column=2, columnspan=4, sticky="w", pady=(10, 0))

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, sticky="ew", pady=(0, 10))

        self.discover_btn = ttk.Button(btn_frame, text="1. Discover Hotels", command=self.start_discovery, width=20)
        self.discover_btn.pack(side="left", padx=(0, 10))

        self.enrich_btn = ttk.Button(btn_frame, text="2. Enrich Data", command=self.start_enrichment, width=20, state="disabled")
        self.enrich_btn.pack(side="left", padx=(0, 10))

        self.stop_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_process, width=10, state="disabled")
        self.stop_btn.pack(side="left", padx=(0, 10))

        self.export_btn = ttk.Button(btn_frame, text="3. Export Excel", command=self.export_to_excel, width=20, state="disabled")
        self.export_btn.pack(side="left")

        # Results table
        results_frame = ttk.LabelFrame(main_frame, text="Hotels", padding="5")
        results_frame.grid(row=4, column=0, sticky="nsew", pady=(0, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        columns = ('org_number', 'legal_name', 'commercial_name', 'address', 'municipality', 'type', 'stars', 'brand', 'phone', 'website', 'status')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=18)

        col_widths = {'org_number': 80, 'legal_name': 180, 'commercial_name': 150, 'address': 200,
                      'municipality': 90, 'type': 60, 'stars': 40, 'brand': 80, 'phone': 100, 'website': 120, 'status': 70}
        for col in columns:
            self.tree.heading(col, text=col.replace('_', ' ').title())
            self.tree.column(col, width=col_widths.get(col, 100), minwidth=40)

        scrollbar_y = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Status
        self.status_var = tk.StringVar(value="Ready. Click 'Discover Hotels' to start.")
        ttk.Label(main_frame, textvariable=self.status_var).grid(row=5, column=0, sticky="w")

        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=6, column=0, sticky="ew", pady=(5, 0))

        self.stats_var = tk.StringVar(value="Hotels: 0 | Enriched: 0 | API calls: 0/300")
        ttk.Label(main_frame, textvariable=self.stats_var, font=('Helvetica', 9)).grid(row=7, column=0, sticky="w", pady=(5, 0))

    def stop_process(self):
        self.is_running = False
        self.status_var.set("Stopping...")

    def start_discovery(self):
        if self.is_running:
            return

        self.is_running = True
        self.hotels = []
        self.clear_tree()

        self.discover_btn.config(state="disabled")
        self.enrich_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress.start()

        thread = threading.Thread(target=self.discover_hotels)
        thread.daemon = True
        thread.start()

    def discover_hotels(self):
        """Discover hotels using Brreg API (free, unlimited)"""
        region = self.region_var.get()
        fylke_prefixes = FYLKE_PREFIX.get(region, [])

        try:
            limit = min(int(self.limit_var.get()), 300)  # Cap at 300 for free Google
        except:
            limit = 300

        # Build NACE code list
        nace_codes = []
        if self.include_hotels.get():
            nace_codes.extend(NACE_CODES["hotels"])
        if self.include_bb.get():
            nace_codes.extend(NACE_CODES["bb"])
        if self.include_camping.get():
            nace_codes.extend(NACE_CODES["camping"])

        if not nace_codes:
            nace_codes = NACE_CODES["hotels"]

        self.update_status(f"Discovering hotels in {region} via Brreg API...")

        seen_orgs = set()

        for nace in nace_codes:
            if not self.is_running or len(self.hotels) >= limit:
                break

            self.update_status(f"Searching NACE {nace}...")

            # Brreg API call - get all, filter by region locally
            page = 0
            while self.is_running and len(self.hotels) < limit:
                url = "https://data.brreg.no/enhetsregisteret/api/enheter"
                params = {
                    'naeringskode': nace,
                    'size': 100,
                    'page': page
                }

                try:
                    response = requests.get(url, params=params, timeout=30)
                    if response.status_code != 200:
                        break

                    data = response.json()
                    companies = data.get('_embedded', {}).get('enheter', [])

                    if not companies:
                        break

                    for company in companies:
                        if not self.is_running or len(self.hotels) >= limit:
                            break

                        org = company.get('organisasjonsnummer', '')
                        if org in seen_orgs:
                            continue

                        # Get address info
                        addr = company.get('forretningsadresse', {}) or company.get('postadresse', {})
                        kommune_nr = addr.get('kommunenummer', '')

                        # Filter by region (fylke prefix)
                        if fylke_prefixes and kommune_nr:
                            if not any(kommune_nr.startswith(prefix) for prefix in fylke_prefixes):
                                continue  # Skip if not in selected region

                        seen_orgs.add(org)

                        # Extract address
                        address_parts = addr.get('adresse', [])
                        address = ', '.join(address_parts) if address_parts else ''
                        postal = f"{addr.get('postnummer', '')} {addr.get('poststed', '')}".strip()
                        full_address = f"{address}, {postal}" if address else postal

                        hotel = {
                            'org_number': org,
                            'legal_name': company.get('navn', ''),
                            'commercial_name': '',
                            'address': full_address,
                            'municipality': addr.get('kommune', ''),
                            'property_type': self.classify_type(company.get('navn', ''), nace),
                            'stars': '',
                            'brand': '',
                            'phone': '',
                            'website': '',
                            'google_rating': '',
                            'status': 'Discovered'
                        }

                        self.hotels.append(hotel)
                        self.root.after(0, lambda h=hotel: self.add_tree_row(h))
                        self.update_stats()

                    page += 1
                    time.sleep(0.3)  # Be nice to API

                except Exception as e:
                    print(f"Brreg error: {e}")
                    break

        self.is_running = False
        self.root.after(0, self.discovery_complete)

    def classify_type(self, name, nace):
        name_lower = name.lower()
        if 'camping' in name_lower or nace == '55.300':
            return 'Camping'
        elif any(x in name_lower for x in ['pensjonat', 'gjestehus', 'b&b']):
            return 'B&B'
        elif 'vandrerhjem' in name_lower or 'hostel' in name_lower:
            return 'Hostel'
        elif 'resort' in name_lower:
            return 'Resort'
        elif 'lodge' in name_lower:
            return 'Lodge'
        return 'Hotel'

    def discovery_complete(self):
        self.progress.stop()
        self.discover_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

        if self.hotels:
            self.enrich_btn.config(state="normal")
            self.export_btn.config(state="normal")

        self.status_var.set(f"Found {len(self.hotels)} hotels. Click 'Enrich Data' to get details from Google.")

    def start_enrichment(self):
        if self.is_running or not self.hotels:
            return

        if self.api_calls >= self.MAX_API_CALLS:
            messagebox.showwarning("API Limit", f"Already used {self.api_calls}/300 free API calls today.")
            return

        self.is_running = True
        self.discover_btn.config(state="disabled")
        self.enrich_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress.start()

        thread = threading.Thread(target=self.enrich_hotels)
        thread.daemon = True
        thread.start()

    def enrich_hotels(self):
        """Enrich with Google Places API (max 300 free/day)"""
        total = len(self.hotels)
        enriched = 0

        for idx, hotel in enumerate(self.hotels):
            if not self.is_running:
                break

            if self.api_calls >= self.MAX_API_CALLS:
                self.update_status(f"Reached free API limit (300/day). Stopping.")
                break

            legal_name = hotel.get('legal_name', '')
            address = hotel.get('address', '')

            self.update_status(f"Enriching {idx + 1}/{total}: {legal_name[:40]}... (API: {self.api_calls}/300)")

            # Google Places lookup
            google_data = self.lookup_google(legal_name, address)

            if google_data:
                hotel['commercial_name'] = google_data.get('name', '')
                hotel['address'] = google_data.get('formatted_address', '') or hotel['address']
                hotel['google_rating'] = google_data.get('rating', '')
                hotel['stars'] = self.rating_to_stars(google_data.get('rating'))
                hotel['status'] = 'Enriched'
                enriched += 1
            else:
                hotel['status'] = 'No match'

            # Detect brand from name
            hotel['brand'] = self.detect_brand(hotel.get('commercial_name') or legal_name)

            # Update UI
            self.root.after(0, lambda h=hotel, i=idx: self.update_tree_row(i, h))
            self.update_stats()

            # Small delay between API calls
            time.sleep(0.3)

        self.is_running = False
        self.root.after(0, self.enrichment_complete)

    def lookup_google(self, name, address):
        """Call Google Places API"""
        if not GOOGLE_PLACES_API_KEY:
            return None

        try:
            # Clean company suffixes
            clean_name = re.sub(r'\b(AS|ANS|DA|ENK|DRIFT|AVD)\b', '', name, flags=re.IGNORECASE).strip()
            query = f"{clean_name} Norway"

            url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json"
            params = {
                'input': query,
                'inputtype': 'textquery',
                'fields': 'name,rating,formatted_address',
                'key': GOOGLE_PLACES_API_KEY
            }

            response = requests.get(url, params=params, timeout=10)
            self.api_calls += 1

            data = response.json()
            print(f"Google API response for '{query}': status={data.get('status')}, candidates={len(data.get('candidates', []))}")

            if data.get('status') == 'OK' and data.get('candidates'):
                return data['candidates'][0]
            else:
                print(f"Google API issue: {data}")

        except Exception as e:
            print(f"Google API error: {e}")

        return None

    def rating_to_stars(self, rating):
        try:
            r = float(rating)
            if r >= 4.5: return '5'
            elif r >= 4.0: return '4'
            elif r >= 3.5: return '3'
            elif r >= 3.0: return '2'
            return '1'
        except:
            return ''

    def detect_brand(self, name):
        brands = {
            'Thon': 'Thon Hotels', 'Scandic': 'Scandic', 'Clarion': 'Nordic Choice',
            'Comfort': 'Nordic Choice', 'Quality': 'Nordic Choice', 'Radisson': 'Radisson',
            'Hilton': 'Hilton', 'Best Western': 'Best Western', 'Smarthotel': 'Smarthotel',
        }
        for key, brand in brands.items():
            if key.upper() in name.upper():
                return brand
        return ''

    def enrichment_complete(self):
        self.progress.stop()
        self.discover_btn.config(state="normal")
        self.enrich_btn.config(state="normal")
        self.export_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

        enriched = sum(1 for h in self.hotels if h.get('status') == 'Enriched')
        self.status_var.set(f"Done! Enriched {enriched}/{len(self.hotels)} hotels. API calls: {self.api_calls}/300")

    def clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def add_tree_row(self, hotel):
        self.tree.insert('', 'end', values=(
            hotel.get('org_number', ''),
            hotel.get('legal_name', '')[:35],
            hotel.get('commercial_name', '')[:30],
            hotel.get('address', '')[:40],
            hotel.get('municipality', ''),
            hotel.get('property_type', ''),
            hotel.get('stars', ''),
            hotel.get('brand', ''),
            hotel.get('phone', ''),
            hotel.get('website', '')[:25] if hotel.get('website') else '',
            hotel.get('status', '')
        ))

    def update_tree_row(self, idx, hotel):
        items = self.tree.get_children()
        if idx < len(items):
            self.tree.item(items[idx], values=(
                hotel.get('org_number', ''),
                hotel.get('legal_name', '')[:35],
                hotel.get('commercial_name', '')[:30],
                hotel.get('address', '')[:40],
                hotel.get('municipality', ''),
                hotel.get('property_type', ''),
                hotel.get('stars', ''),
                hotel.get('brand', ''),
                hotel.get('phone', ''),
                hotel.get('website', '')[:25] if hotel.get('website') else '',
                hotel.get('status', '')
            ))

    def update_status(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))

    def update_stats(self):
        enriched = sum(1 for h in self.hotels if h.get('status') == 'Enriched')
        self.root.after(0, lambda: self.stats_var.set(f"Hotels: {len(self.hotels)} | Enriched: {enriched} | API calls: {self.api_calls}/300"))

    def export_to_excel(self):
        if not self.hotels:
            messagebox.showwarning("No Data", "No hotels to export.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Norway_Hotels_{self.region_var.get()}_{timestamp}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=filename
        )

        if not filepath:
            return

        try:
            df = pd.DataFrame(self.hotels)
            df['export_date'] = datetime.now().strftime('%Y-%m-%d')

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Hotels', index=False)

                ws = writer.sheets['Hotels']
                from openpyxl.styles import Font, PatternFill

                for cell in ws[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

                ws.freeze_panes = 'A2'

            messagebox.showinfo("Success", f"Exported {len(self.hotels)} hotels to:\n{filepath}")
            os.startfile(os.path.dirname(filepath))

        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")


def main():
    root = tk.Tk()
    ttk.Style().theme_use('clam')
    HotelScraperApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
